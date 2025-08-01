"""
Excel ID-Based Comparator
=========================

Script Name: Excel ID-Based Comparator
Script URI: https://github.com/thomasgerdes/excel-id-comparator
Description: A robust Python tool for comparing Excel files based on unique identifiers, highlighting actual data changes rather than positional differences
Version: 1.0.0
Author: Thomas Gerdes
Author URI: https://thomasgerdes.de
License: MIT
License URI: https://opensource.org/licenses/MIT
Requires Python: 3.7+
Compatible with: Windows, macOS, Linux, Google Colab, Jupyter Notebook
Dependencies: pandas, openpyxl, numpy

Features:
- ID-based comparison (not position-based)
- Automatic sheet and ID column detection
- Configurable parameters
- Robust error handling
- Clean visual output with change highlighting
- Separate tracking of added, modified, and deleted records
- Cross-platform compatibility
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import os
from datetime import datetime
import copy
import argparse
from typing import Dict, List, Tuple, Optional, Any

class ExcelIDComparator:
    """
    A class for comparing Excel files based on unique identifiers.
    
    This comparator focuses on content changes rather than positional changes,
    making it ideal for datasets where rows may be added, removed, or reordered.
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize the Excel ID Comparator.
        
        Args:
            config (dict, optional): Configuration dictionary with the following keys:
                - sheet_name (str): Name of the sheet to compare (auto-detect if None)
                - id_column (str): Column letter for IDs (auto-detect if None)  
                - id_column_index (int): Column index for IDs (0-based, auto-detect if None)
                - case_sensitive (bool): Whether comparisons are case-sensitive (default: True)
                - ignore_empty_cells (bool): Whether to ignore empty cells in comparison (default: True)
        """
        self.config = config or {}
        
        # Styling for different types of changes
        self.styles = {
            'changed': {
                'font': Font(color="CC0000", bold=True),  # Red for changes
                'fill': None  # No background for clean look
            },
            'new': {
                'font': Font(color="008000", bold=True),  # Green for new records
                'fill': PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
            },
            'deleted': {
                'font': Font(color="FF8C00", bold=True),  # Orange for deleted records
                'fill': PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
            }
        }
        
        # Statistics tracking
        self.stats = {
            'total_ids_file1': 0,
            'total_ids_file2': 0,
            'modified_ids': 0,
            'new_ids': 0,
            'deleted_ids': 0,
            'unchanged_ids': 0,
            'processing_errors': 0
        }
    
    def detect_sheet_and_id_column(self, workbook: Any, filename: str) -> Tuple[Optional[str], Optional[int]]:
        """
        Automatically detect the best sheet and ID column to use.
        
        Returns:
            Tuple of (sheet_name, id_column_index) or (None, None) if detection fails
        """
        print(f"🔍 Auto-detecting structure in {filename}...")
        
        # If sheet name is configured, use it
        if self.config.get('sheet_name'):
            sheet_name = self.config['sheet_name']
            if sheet_name in workbook.sheetnames:
                print(f"   ✅ Using configured sheet: {sheet_name}")
            else:
                print(f"   ❌ Configured sheet '{sheet_name}' not found")
                print(f"   📋 Available sheets: {workbook.sheetnames}")
                print(f"   🔄 Falling back to auto-detection...")
                # Clear the configured sheet name and auto-detect
                self.config['sheet_name'] = None
        
        if not self.config.get('sheet_name'):
            # Auto-detect: prefer first sheet with data, avoid common metadata sheets
            skip_sheets = ['about', 'readme', 'info', 'metadata', 'codebook']
            sheet_name = None
            
            for name in workbook.sheetnames:
                if name.lower() not in skip_sheets:
                    ws = workbook[name]
                    if ws.max_row > 1:  # Has data beyond header
                        sheet_name = name
                        print(f"   ✅ Auto-detected sheet: {sheet_name}")
                        break
            
            if not sheet_name:
                sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
                print(f"   ⚠️ Fallback to first sheet: {sheet_name}")
        else:
            sheet_name = self.config['sheet_name']
        
        if not sheet_name:
            return None, None
        
        # Detect ID column
        ws = workbook[sheet_name]
        id_column_index = None
        
        if self.config.get('id_column_index') is not None:
            id_column_index = self.config['id_column_index']
            print(f"   ✅ Using configured ID column index: {id_column_index}")
        elif self.config.get('id_column'):
            # Convert column letter to index
            col_letter = self.config['id_column'].upper()
            id_column_index = ord(col_letter) - ord('A')
            print(f"   ✅ Using configured ID column: {col_letter} (index {id_column_index})")
        else:
            # Auto-detect ID column: look for columns with "ID" in header or first column
            print("   🔍 Auto-detecting ID column...")
            
            # Check first row for headers containing "id"
            for col in range(min(10, ws.max_column)):  # Check first 10 columns
                cell = ws.cell(1, col + 1)
                if cell.value and 'id' in str(cell.value).lower():
                    id_column_index = col
                    print(f"   ✅ Found ID column by header: {chr(65 + col)} ('{cell.value}')")
                    break
            
            # Fallback to column A
            if id_column_index is None:
                id_column_index = 0
                print(f"   ⚠️ Fallback to column A as ID column")
        
        return sheet_name, id_column_index
    
    def extract_data_from_file(self, filepath: str) -> Optional[Dict]:
        """
        Extract ID-based data from an Excel file.
        
        Args:
            filepath (str): Path to the Excel file
            
        Returns:
            Dictionary containing extracted data or None if extraction fails
        """
        try:
            print(f"📖 Reading {os.path.basename(filepath)}...")
            workbook = load_workbook(filepath, data_only=True)
            
            # Detect sheet and ID column
            sheet_name, id_column_index = self.detect_sheet_and_id_column(workbook, filepath)
            if not sheet_name or id_column_index is None:
                print(f"❌ Could not detect structure in {filepath}")
                return None
            
            ws = workbook[sheet_name]
            print(f"   📊 Sheet size: {ws.max_row} rows × {ws.max_column} columns")
            
            # Extract headers
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(1, col)
                header_value = str(cell.value).strip() if cell.value else f"Column_{col}"
                headers.append(header_value)
            
            print(f"   📋 Found {len(headers)} columns")
            print(f"   🔍 ID column: {chr(65 + id_column_index)} ('{headers[id_column_index]}')")
            
            # Extract data rows
            id_data = {}
            processed_rows = 0
            errors = 0
            
            for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                try:
                    # Get ID value
                    id_cell = ws.cell(row_idx, id_column_index + 1)
                    if not id_cell.value:
                        continue
                    
                    id_value = str(id_cell.value).strip()
                    if not id_value:
                        continue
                    
                    # Handle case sensitivity
                    if not self.config.get('case_sensitive', True):
                        id_value = id_value.lower()
                    
                    # Extract row data
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        try:
                            cell = ws.cell(row_idx, col_idx + 1)
                            value = cell.value if cell.value is not None else ''
                            
                            # Clean and convert value
                            if value == '' and self.config.get('ignore_empty_cells', True):
                                clean_value = ''
                            else:
                                clean_value = str(value).strip()
                                if not self.config.get('case_sensitive', True):
                                    clean_value = clean_value.lower()
                            
                            row_data[header] = clean_value
                            
                        except Exception as e:
                            row_data[header] = ''
                            errors += 1
                    
                    # Store the data
                    if id_value in id_data:
                        print(f"   ⚠️ Duplicate ID found: {id_value} (keeping first occurrence)")
                    else:
                        id_data[id_value] = {
                            'row_num': row_idx,
                            'data': row_data
                        }
                        processed_rows += 1
                
                except Exception as e:
                    errors += 1
                    continue
            
            print(f"   ✅ Processed {processed_rows} records")
            if errors > 0:
                print(f"   ⚠️ {errors} cell reading errors (values set to empty)")
            
            # Show sample IDs
            if id_data:
                sample_ids = list(id_data.keys())[:5]
                print(f"   🔍 Sample IDs: {sample_ids}")
            
            return {
                'id_data': id_data,
                'headers': headers,
                'sheet_name': sheet_name,
                'id_column_index': id_column_index,
                'filepath': filepath,
                'processed_rows': processed_rows
            }
            
        except Exception as e:
            print(f"❌ Error reading {filepath}: {str(e)}")
            return None
    
    def compare_datasets(self, data1: Dict, data2: Dict) -> Dict:
        """
        Compare two datasets and identify changes.
        
        Args:
            data1: Data from first file (reference)
            data2: Data from second file (comparison)
            
        Returns:
            Dictionary containing comparison results
        """
        print("\n🔍 COMPARING DATASETS...")
        print("=" * 50)
        
        id_data1 = data1['id_data']
        id_data2 = data2['id_data']
        
        self.stats['total_ids_file1'] = len(id_data1)
        self.stats['total_ids_file2'] = len(id_data2)
        
        print(f"📊 File 1: {len(id_data1)} records")
        print(f"📊 File 2: {len(id_data2)} records")
        
        all_ids = set(id_data1.keys()) | set(id_data2.keys())
        print(f"📊 Total unique IDs: {len(all_ids)}")
        
        comparison_results = {
            'modified_records': {},
            'new_records': {},
            'deleted_records': {},
            'unchanged_records': set()
        }
        
        # Process each ID
        for id_value in all_ids:
            try:
                in_file1 = id_value in id_data1
                in_file2 = id_value in id_data2
                
                if in_file1 and in_file2:
                    # Compare records that exist in both files
                    record1 = id_data1[id_value]['data']
                    record2 = id_data2[id_value]['data']
                    
                    # Find field-level changes
                    field_changes = {}
                    all_fields = set(record1.keys()) | set(record2.keys())
                    
                    for field in all_fields:
                        val1 = record1.get(field, '')
                        val2 = record2.get(field, '')
                        
                        if val1 != val2:
                            field_changes[field] = {
                                'old_value': val1,
                                'new_value': val2
                            }
                    
                    if field_changes:
                        comparison_results['modified_records'][id_value] = {
                            'row_num': id_data2[id_value]['row_num'],
                            'changes': field_changes,
                            'record_data': record2
                        }
                        self.stats['modified_ids'] += 1
                    else:
                        comparison_results['unchanged_records'].add(id_value)
                        self.stats['unchanged_ids'] += 1
                
                elif in_file2 and not in_file1:
                    # New record (only in file 2)
                    comparison_results['new_records'][id_value] = {
                        'row_num': id_data2[id_value]['row_num'],
                        'record_data': id_data2[id_value]['data']
                    }
                    self.stats['new_ids'] += 1
                
                elif in_file1 and not in_file2:
                    # Deleted record (only in file 1)
                    comparison_results['deleted_records'][id_value] = {
                        'row_num': id_data1[id_value]['row_num'],
                        'record_data': id_data1[id_value]['data']
                    }
                    self.stats['deleted_ids'] += 1
                    
            except Exception as e:
                print(f"   ⚠️ Error processing ID {id_value}: {e}")
                self.stats['processing_errors'] += 1
                continue
        
        # Print summary
        print(f"\n📊 COMPARISON SUMMARY:")
        print(f"🔴 Modified records: {self.stats['modified_ids']}")
        print(f"🟢 New records: {self.stats['new_ids']}")
        print(f"🟠 Deleted records: {self.stats['deleted_ids']}")
        print(f"⚪ Unchanged records: {self.stats['unchanged_ids']}")
        
        if self.stats['processing_errors'] > 0:
            print(f"⚠️ Processing errors: {self.stats['processing_errors']}")
        
        return comparison_results
    
    def create_comparison_report(self, data1: Dict, data2: Dict, comparison_results: Dict, output_path: str) -> str:
        """
        Create an Excel report showing the comparison results.
        
        Args:
            data1: Data from first file
            data2: Data from second file  
            comparison_results: Results from comparison
            output_path: Path for output file
            
        Returns:
            Path to the created report file
        """
        print(f"\n📊 CREATING COMPARISON REPORT...")
        print("=" * 50)
        
        try:
            # Create new workbook based on file 2
            original_wb = load_workbook(data2['filepath'])
            report_wb = copy.copy(original_wb)
            
            # Get the main sheet
            main_sheet = report_wb[data2['sheet_name']]
            
            # Apply change markings
            self._apply_change_markings(main_sheet, comparison_results, data2)
            
            # Add summary sheet
            self._add_summary_sheet(report_wb, data1, data2, comparison_results)
            
            # Add deleted records sheet if any
            if comparison_results['deleted_records']:
                self._add_deleted_records_sheet(report_wb, data1, comparison_results)
            
            # Save the report
            report_wb.save(output_path)
            print(f"✅ Report saved: {output_path}")
            
            return output_path
            
        except Exception as e:
            print(f"❌ Error creating report: {e}")
            raise
    
    def _apply_change_markings(self, worksheet, comparison_results: Dict, data2: Dict):
        """Apply visual change markings to the worksheet."""
        print("🎨 Applying change markings...")
        
        marked_changes = 0
        marked_new = 0
        
        # Mark modified records
        for id_value, info in comparison_results['modified_records'].items():
            try:
                row_num = info['row_num']
                changes = info['changes']
                headers = data2['headers']
                
                # Mark only changed fields
                for col_idx, header in enumerate(headers):
                    if header in changes:
                        cell = worksheet.cell(row_num, col_idx + 1)
                        cell.font = self.styles['changed']['font']
                        
                        # Add comment with old value
                        old_value = changes[header]['old_value']
                        if old_value:
                            comment = Comment(f"Previous: {old_value}", "ID-Comparator")
                            cell.comment = comment
                        
                        marked_changes += 1
                        
            except Exception as e:
                print(f"   ⚠️ Error marking changes for ID {id_value}: {e}")
                continue
        
        # Mark new records
        for id_value, info in comparison_results['new_records'].items():
            try:
                row_num = info['row_num']
                headers = data2['headers']
                
                # Mark entire row as new
                for col_idx in range(len(headers)):
                    cell = worksheet.cell(row_num, col_idx + 1)
                    cell.font = self.styles['new']['font']
                    if self.styles['new']['fill']:
                        cell.fill = self.styles['new']['fill']
                    
                    # Add comment to ID column
                    if col_idx == data2['id_column_index']:
                        comment = Comment(f"New record: {id_value}", "ID-Comparator")
                        cell.comment = comment
                
                marked_new += 1
                
            except Exception as e:
                print(f"   ⚠️ Error marking new record {id_value}: {e}")
                continue
        
        print(f"   ✅ Marked {marked_changes} changed fields")
        print(f"   ✅ Marked {marked_new} new records")
    
    def _add_summary_sheet(self, workbook, data1: Dict, data2: Dict, comparison_results: Dict):
        """Add a summary sheet to the workbook."""
        summary_sheet = workbook.create_sheet("📊 Comparison Summary", 0)
        
        summary_data = [
            ["EXCEL ID-BASED COMPARISON REPORT"],
            [""],
            ["📄 Reference File:", os.path.basename(data1['filepath'])],
            ["📄 Comparison File:", os.path.basename(data2['filepath'])],
            ["🎯 Compared Sheet:", data2['sheet_name']],
            ["🔑 ID Column:", f"{chr(65 + data2['id_column_index'])} ({data2['headers'][data2['id_column_index']]})"],
            ["🕐 Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [""],
            ["📊 COMPARISON RESULTS:"],
            ["🔴 Modified Records:", self.stats['modified_ids']],
            ["🟢 New Records:", self.stats['new_ids']],
            ["🟠 Deleted Records:", self.stats['deleted_ids']],
            ["⚪ Unchanged Records:", self.stats['unchanged_ids']],
            ["📈 Total Changes:", self.stats['modified_ids'] + self.stats['new_ids'] + self.stats['deleted_ids']],
            [""],
            ["🎨 LEGEND:"],
            ["🔴 Red Text = Modified values (comments show previous values)"],
            ["🟢 Green Text + Background = New records"],
            ["🟠 See 'Deleted Records' sheet = Removed records"],
            ["⚪ Normal Text = Unchanged values"],
            [""],
            ["⚙️ CONFIGURATION:"],
            ["Case Sensitive:", self.config.get('case_sensitive', True)],
            ["Ignore Empty Cells:", self.config.get('ignore_empty_cells', True)],
        ]
        
        for row_data in summary_data:
            summary_sheet.append(row_data)
        
        # Format the summary sheet
        summary_sheet['A1'].font = Font(size=16, bold=True, color="000080")
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 40
    
    def _add_deleted_records_sheet(self, workbook, data1: Dict, comparison_results: Dict):
        """Add a sheet showing deleted records."""
        deleted_sheet = workbook.create_sheet("🗑️ Deleted Records")
        
        headers = ["ID"] + [h for h in data1['headers'] if h != data1['headers'][data1['id_column_index']]]
        
        # Add title and headers
        deleted_sheet.append([f"Records deleted from {os.path.basename(data1['filepath'])}"])
        deleted_sheet.append([])
        deleted_sheet.append(headers)
        
        # Add deleted records
        for id_value, info in comparison_results['deleted_records'].items():
            try:
                row_data = [id_value]
                record_data = info['record_data']
                
                for header in headers[1:]:  # Skip ID column as it's already added
                    row_data.append(record_data.get(header, ''))
                
                deleted_sheet.append(row_data)
                
            except Exception as e:
                print(f"   ⚠️ Error adding deleted record {id_value}: {e}")
                continue
        
        # Format deleted records
        for row in range(4, deleted_sheet.max_row + 1):  # Start from data rows
            for col in range(1, deleted_sheet.max_column + 1):
                cell = deleted_sheet.cell(row, col)
                cell.font = self.styles['deleted']['font']
                if self.styles['deleted']['fill']:
                    cell.fill = self.styles['deleted']['fill']
    
    def compare_files(self, file1_path: str, file2_path: str, output_path: Optional[str] = None) -> str:
        """
        Main method to compare two Excel files.
        
        Args:
            file1_path: Path to the reference file
            file2_path: Path to the comparison file
            output_path: Path for output file (auto-generated if None)
            
        Returns:
            Path to the generated comparison report
        """
        print("🔍 EXCEL ID-BASED COMPARISON")
        print("=" * 60)
        print("Generic tool for comparing Excel files based on unique IDs")
        print()
        
        # Extract data from both files
        data1 = self.extract_data_from_file(file1_path)
        if not data1:
            raise ValueError(f"Could not extract data from {file1_path}")
        
        data2 = self.extract_data_from_file(file2_path)
        if not data2:
            raise ValueError(f"Could not extract data from {file2_path}")
        
        # Validate compatibility
        if data1['headers'] != data2['headers']:
            print("⚠️ WARNING: Files have different column structures")
            print(f"   File 1 columns: {len(data1['headers'])}")  
            print(f"   File 2 columns: {len(data2['headers'])}")
        
        # Perform comparison
        comparison_results = self.compare_datasets(data1, data2)
        
        # Generate output path if not provided
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(os.path.basename(file2_path))[0]
            output_path = f"comparison_{base_name}_{timestamp}.xlsx"
        
        # Create report
        report_path = self.create_comparison_report(data1, data2, comparison_results, output_path)
        
        # Show final summary
        self._show_final_summary(file1_path, file2_path, report_path)
        
        return report_path
    
    def _show_final_summary(self, file1_path: str, file2_path: str, report_path: str):
        """Show final summary of the comparison."""
        print("\n🎉 COMPARISON COMPLETED!")
        print("=" * 50)
        
        total_changes = self.stats['modified_ids'] + self.stats['new_ids'] + self.stats['deleted_ids']
        
        if total_changes == 0:
            print("✨ No differences found - files are identical!")
        else:
            print(f"📊 Found {total_changes} total changes:")
            print(f"   🔴 {self.stats['modified_ids']} modified records")
            print(f"   🟢 {self.stats['new_ids']} new records")
            print(f"   🟠 {self.stats['deleted_ids']} deleted records")
        
        print(f"\n📁 Files processed:")
        print(f"   📄 Reference: {os.path.basename(file1_path)}")
        print(f"   📄 Comparison: {os.path.basename(file2_path)}")
        print(f"   📊 Report: {os.path.basename(report_path)}")
        
        print(f"\n💡 Open '{os.path.basename(report_path)}' to see highlighted changes!")


def run_colab_version():
    """Google Colab-specific version with file upload."""
    print("🔍 EXCEL ID-BASED COMPARATOR - GOOGLE COLAB")
    print("=" * 50)
    print("Upload Excel files and compare based on unique IDs")
    print()
    
    try:
        from google.colab import files
    except ImportError:
        print("❌ This function requires Google Colab")
        return
    
    # Upload first file
    print("📁 Upload FIRST file (reference/original):")
    uploaded1 = files.upload()
    if not uploaded1:
        print("❌ No file uploaded!")
        return
    
    file1_name = list(uploaded1.keys())[0]
    print(f"✅ Reference: {file1_name}")
    print()
    
    # Upload second file
    print("📁 Upload SECOND file (comparison/new):")
    uploaded2 = files.upload()
    if not uploaded2:
        print("❌ No file uploaded!")
        return
        
    file2_name = list(uploaded2.keys())[0]
    print(f"✅ Comparison: {file2_name}")
    print()
    
    print("⚙️ CONFIGURATION (press Enter for defaults):")
    
    # Show available sheets from first file for user guidance
    try:
        from openpyxl import load_workbook
        wb_temp = load_workbook(file1_name, data_only=True)
        available_sheets = wb_temp.sheetnames
        print(f"   Available sheets: {', '.join(available_sheets)}")
        wb_temp.close()
    except:
        print("   Available sheets: Could not detect")
    
    sheet_name = input("Sheet name (auto-detect if empty): ").strip()
    
    print("   ID column: Auto-detect if empty (A, B, C, etc.)")
    id_column = input("ID column: ").strip().upper()
    
    print("   Case sensitive: No if empty")
    case_sensitive = input("Case sensitive (y/N): ").strip().lower()
    
    # Build configuration
    config = {
        'case_sensitive': case_sensitive == 'y',
        'ignore_empty_cells': True
    }
    
    if sheet_name:
        config['sheet_name'] = sheet_name
    
    if id_column:
        config['id_column'] = id_column
    
    print("\n🔄 Processing...")
    
    try:
        # Create comparator and run comparison
        comparator = ExcelIDComparator(config)
        report_path = comparator.compare_files(file1_name, file2_name)
        
        # Download the result
        print(f"\n✅ Comparison completed!")
        print(f"📊 Report: {report_path}")
        print("\n💾 Downloading...")
        files.download(report_path)
        
        print("\n🎉 DONE! Open the downloaded file to see changes.")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()


def compare_excel_colab():
    """
    Convenience function for Google Colab users.
    Simply call this function to start the comparison process.
    """
    run_colab_version()


def main():
    """Command line interface for the Excel ID Comparator."""
    parser = argparse.ArgumentParser(
        description="Compare Excel files based on unique IDs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_id_comparator.py file1.xlsx file2.xlsx
  python excel_id_comparator.py file1.xlsx file2.xlsx --sheet "Data" --id-column "A"
  python excel_id_comparator.py file1.xlsx file2.xlsx --output comparison_report.xlsx
        """
    )
    
    parser.add_argument("file1", help="Path to the reference Excel file")
    parser.add_argument("file2", help="Path to the comparison Excel file")
    parser.add_argument("-o", "--output", help="Output file path (auto-generated if not specified)")
    parser.add_argument("-s", "--sheet", help="Sheet name to compare (auto-detect if not specified)")
    parser.add_argument("-c", "--id-column", help="ID column letter (A, B, C, etc.) - auto-detect if not specified")
    parser.add_argument("--case-insensitive", action="store_true", help="Perform case-insensitive comparison")
    parser.add_argument("--include-empty", action="store_true", help="Include empty cells in comparison")
    
    args = parser.parse_args()
    
    # Validate input files
    if not os.path.exists(args.file1):
        print(f"❌ Error: File not found: {args.file1}")
        return 1
    
    if not os.path.exists(args.file2):
        print(f"❌ Error: File not found: {args.file2}")
        return 1
    
    # Build configuration
    config = {
        'case_sensitive': not args.case_insensitive,
        'ignore_empty_cells': not args.include_empty
    }
    
    if args.sheet:
        config['sheet_name'] = args.sheet
    
    if args.id_column:
        config['id_column'] = args.id_column.upper()
    
    # Create comparator and run comparison
    try:
        comparator = ExcelIDComparator(config)
        report_path = comparator.compare_files(args.file1, args.file2, args.output)
        print(f"\n✅ Success! Report generated: {report_path}")
        return 0
        
    except Exception as e:
        print(f"\n❌ Error during comparison: {e}")
        return 1


if __name__ == "__main__":
    # Check if running in Colab and handle accordingly
    try:
        import google.colab
        # In Colab, run the interactive version
        run_colab_version()
    except ImportError:
        # Regular command line execution
        exit(main())
